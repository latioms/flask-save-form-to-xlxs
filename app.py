import random
from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
from wtforms import Form, StringField, validators

app = Flask(__name__)

# Création du formulaire
class DataForm(Form):
  name = StringField('Nom', [validators.Length(min=1, max=50)])
  email = StringField('Email', [validators.Length(min=6, max=50)])

# Route principale qui affiche le formulaire
@app.route('/')
def index():
  form = DataForm(request.form)

  # Chargement du fichier Excel et récupération des données
  wb = load_workbook('data.xlsx')
  sheet = wb.active
  data = []
  i = 1
  for row in sheet.iter_rows(min_row=2):
    data.append({
      'number': i,
      'name': row[0].value,
      'email': row[1].value,
    })
    i +=1

  return render_template('index.html', form=form, data=data)

# Route qui traite les données du formulaire
@app.route('/process', methods=['POST'])
def process():
  form = DataForm(request.form)
  if request.method == 'POST' and form.validate():
    # Récupération des données du formulaire
    name = form.name.data
    email = form.email.data

    # Chargement du fichier Excel existant (s'il existe) ou création d'un nouveau fichier
    try:
      wb = load_workbook('data.xlsx')
      sheet = wb.active
    except FileNotFoundError:
      wb = Workbook()
      sheet = wb.active
      sheet['A1'] = 'Nom'
      sheet['B1'] = 'Email'

    # Enregistrement de la nouvelle entrée dans le fichier Excel
    sheet.append([name, email])
    wb.save('data.xlsx')

    return redirect('/success')
  return render_template('index.html', form=form)
  
# Route de confirmation
@app.route('/success')
def success():
  return 'Les données ont été enregistrées avec succès !'



#  Route de la lotterie qui sélectionne au hasard une entrée dans la liste
@app.route('/lotterie')
def lotterie():
  # Chargement du fichier Excel et récupération des données
  wb = load_workbook('data.xlsx')
  sheet = wb.active
  data = []
  for row in sheet.iter_rows(min_row=2):
    data.append({
      'name': row[0].value,
      'email': row[1].value,
    })

  return render_template('lotterie.html', data=data)

# Route qui traite le clic sur le bouton "Gagnant"
@app.route('/lotterie/gagnant')
def gagnant():
  # Chargement du fichier Excel et récupération des données
  wb = load_workbook('data.xlsx')
  sheet = wb.active
  data = []
  for row in sheet.iter_rows(min_row=2):
    data.append({
      'name': row[0].value,
      'email': row[1].value,
    })

  # Sélection au hasard d'une entrée dans la liste
  selected = random.choice(data)

  return render_template('lotterie.html', selected=selected)

if __name__ == '__main__':
  app.run()
